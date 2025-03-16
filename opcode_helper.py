# import pandas lib as pd
import openpyxl
import os

dir_path = os.path.dirname(os.path.realpath(__file__))
Assembled_file = os.path.join(dir_path, 'Architecture.xlsx')
# read by default 1st sheet of an excel file
dataframe = openpyxl.load_workbook(Assembled_file)

dataframe1 = dataframe.active

"""
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
"""

opcodes = []
for col in dataframe1.iter_cols(4, 4):
    for row in range(1, dataframe1.max_row):
        opcodes.append(col[row].value)

for i in range(len(opcodes)):
    opcodes[i] = f'"{opcodes[i]}"'

print(f"[{', '.join(opcodes)}]")